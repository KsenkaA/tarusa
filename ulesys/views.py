import calendar
import datetime as dt
import mimetypes
import os
from datetime import datetime
from wsgiref.util import FileWrapper
from openpyxl.styles import Border, Side, colors, Font, PatternFill
import openpyxl
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.http import StreamingHttpResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from django.shortcuts import render
from .forms import OrderForm, WholeIncomeForm, PartOfIncomeForm, PartOfExpenseForm, CreateExpenseForm
from .models import Order, Partner, PartOfIncome, WholeIncome, WholeExpense, Service, PartOfExpense, \
    PartOfLilExpense, PayDay, GetPayedDay
from django.contrib.auth.models import User


@login_required(login_url='/login/')
def change_order_expense(request, expense_id):
    if request.user.job.admin:
        service_ids = expense_id.split(" ")
        expense = PartOfExpense.objects.filter(service__order__number=service_ids[1])[int(service_ids[0]) - 1]
        services = Service.objects.filter(order__number=service_ids[1])
        if request.method == "POST":
            order = expense.service.order
            create_order_invoice(order)
            try:
                if request.POST['delete'] == 'on':
                    expense.delete()
                    return redirect('/order/' + service_ids[1] + '/')
            except:
                pass
            expense.name = request.POST['name']
            expense.money = request.POST['money']
            expense.currency = request.POST['currency']
            expense.partner = Partner.objects.get(name=request.POST['partner'])
            expense.service = Service.objects.get(order__number=service_ids[1], name=request.POST['service'])
            expense.comments = request.POST['comment']
            expense.save()
            sum = 0
            for exp in expense.expenses.all():
                sum += int(exp.money)
            if sum >= int(expense.money):
                expense.payed = True
            else:
                expense.payed = False
            expense.save()
            return redirect('/order/' + service_ids[1] + '/')
        return render(request, 'change_order_expense.html',
                      {'expense': expense, 'partners': Partner.objects.all(), 'services': services})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_service(request, service_id):
    if request.user.job.admin:
        service_ids = service_id.split(" ")
        service = Service.objects.filter(order__number=service_ids[1])[int(service_ids[0]) - 1]
        if request.method == "POST":
            order = service.order
            create_order_invoice(order)
            try:
                if request.POST['delete'] == 'on':
                    service.delete()
                    return redirect('/order/' + service_ids[1] + '/')
            except:
                pass
            service.name = request.POST['name']
            service.real_price = request.POST['price']
            service.price_for_buyer = request.POST['price_b']
            service.save()

            return redirect('/order/' + service_ids[1] + '/')
        return render(request, 'change_service.html',
                      {'service': service})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_order_m(request, order_id):
    if request.user.job.manager:
        order = Order.objects.get(number=order_id)
        if request.method == "POST":
            create_order_invoice(order)
            order.created = request.POST['date']
            order.tour_starts = request.POST['date_start']
            manager = request.POST['manager'].split()
            order.manager = User.objects.get(first_name=manager[0], last_name=manager[1], job__manager=True)
            order.ref = request.POST['ref']
            order.another_manager = request.POST['another_manager']
            order.last_name = request.POST['last_name']
            order.end_month = request.POST['end_month']
            order.comments = request.POST['comment']
            order.save()

            return redirect('/order/' + str(order_id) + '/')
        date = str(order.created)
        date_start = str(order.tour_starts)
        managers = User.objects.filter(job__manager=True)
        return render(request, 'change_order.html',
                      {'order': order, 'partners': Partner.objects.all(), 'date': date, 'date_start': date_start,
                       'managers': managers})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_order(request, order_id):
    if request.user.job.admin:
        order = Order.objects.get(number=order_id)
        if request.method == "POST":
            create_order_invoice(order)
            try:
                if request.POST['delete'] == 'on':
                    order.delete()
                    return redirect('/search/')
            except:
                pass
            order.created = request.POST['date']
            order.tour_starts = request.POST['date_start']
            manager = request.POST['manager'].split()
            order.manager = User.objects.get(first_name=manager[0], last_name=manager[1], job__manager=True)
            order.ref = request.POST['ref']
            order.another_manager = request.POST['another_manager']
            order.last_name = request.POST['last_name']
            order.end_month = request.POST['end_month']
            order.comments = request.POST['comment']
            order.save()
            return redirect('/order/' + str(order_id) + '/')
        date = str(order.created)
        date_start = str(order.tour_starts)
        managers = User.objects.filter(job__manager=True)
        return render(request, 'change_order.html',
                      {'order': order, 'partners': Partner.objects.all(), 'date': date, 'date_start': date_start,
                       'managers': managers})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_expense_f(request, expense_id):
    if request.user.job.fin_manager:
        expense = WholeExpense.objects.all()[::-1][int(expense_id) - 1]
        if request.method == "POST":
            try:
                expense.file = request.FILES['file']
            except:
                pass
            expense.created = request.POST['date']
            expense.partner = Partner.objects.get(name=request.POST['partner'])
            expense.way_to_pay = request.POST['way']
            expense.money = int(request.POST['money'])
            expense.currency = request.POST['currency']
            expense.comments = request.POST['comment']
            expense.save()
            return redirect('/expenses/')
        date = str(expense.created)
        return render(request, 'change_expense_f.html',
                      {'expense': expense, 'partners': Partner.objects.all(), 'date': date})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_expense(request, expense_id):
    if request.user.job.admin:
        expense = WholeExpense.objects.all()[::-1][int(expense_id) - 1]
        if request.method == "POST":
            try:
                if request.POST['delete'] == 'on':
                    expense.delete()
                    return redirect('/expenses/')
            except:
                pass
            try:
                expense.file = request.FILES['file']
            except:
                pass
            expense.created = request.POST['date']
            expense.partner = Partner.objects.get(name=request.POST['partner'])
            expense.way_to_pay = request.POST['way']
            expense.money = int(request.POST['money'])
            expense.currency = request.POST['currency']
            expense.comments = request.POST['comment']
            expense.save()
            return redirect('/expenses/')
        date = str(expense.created)
        return render(request, 'change_expense.html',
                      {'expense': expense, 'partners': Partner.objects.all(), 'date': date})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_expense_m(request, expense_id):
    if request.user.job.manager:
        expense = WholeExpense.objects.all()[::-1][int(expense_id) - 1]
        if request.method == "POST":
            try:
                expense.file = request.FILES['file']
            except:
                pass
            expense.created = request.POST['date']
            expense.way_to_pay = request.POST['way']
            expense.money = int(request.POST['money'])
            expense.currency = request.POST['currency']
            expense.comments = request.POST['comment']
            expense.save()
            return redirect('/expenses/')
        date = str(expense.created)
        return render(request, 'change_expense_m.html',
                      {'expense': expense, 'partners': Partner.objects.all(), 'date': date})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_part_expense(request, expense_id):
    if request.user.job.admin:
        ids = expense_id.split(".")
        id_w = int(ids[1])
        id_p = int(ids[0])
        flag = True
        try:
            par = ids[2]
        except:
            flag = False
        print(id_w, id_p)
        if flag:
            whole_expense = WholeExpense.objects.get(partner__name=par, pk=id_w)
        else:
            whole_expense = WholeExpense.objects.all()[::-1][id_w - 1]
        expense = PartOfLilExpense.objects.filter(whole_expense=whole_expense)[id_p - 1]
        print(expense)
        part = expense.part_of_expense
        if request.method == "POST":
            order = part.service.order
            create_order_invoice(order)
            try:
                if request.POST['delete'] == 'on':
                    expense.delete()
                    sum = 0
                    for exp in part.expenses.all():
                        sum += int(exp.money)
                    print(sum)
                    print(part.expenses.all())
                    if sum >= part.money:
                        print("!!!!!!!!!!!")
                        part.payed = True
                        part.left_to_pay =  0
                    else:
                        print("123456")
                        part.payed = False
                        part.left_to_pay = part.money - sum
                    part.save()
                    if flag:
                        return redirect('/check/')
                    else:
                        return redirect('/expenses/')
            except:
                pass
            service = request.POST['service'].split('_')
            serv_part = service[1]
            serv = service[0]

            order = Order.objects.get(number=int(request.POST['number']))
            service = Service.objects.get(name=serv, order=order)
            expense.part_of_expense = PartOfExpense.objects.get(service=service, name=serv_part)
            expense.money = int(request.POST['money'])
            expense.currency = request.POST['currency']
            expense.comments = request.POST['comment']
            expense.save()
            sum = 0
            for exp in part.expenses.all():
                sum += int(exp.money)
            print(sum)
            print(part.expenses.all())
            if sum >= part.money:
                print("!!!!!!!!!!!")
                part.payed = True
            else:
                print("123456")
                part.payed = False
            part.save()
            return redirect('/expenses/')
        orders = Order.objects.all()
        services = Service.objects.filter(order=expense.part_of_expense.service.order)
        return render(request, 'change_part_expense.html', {'expense': expense, 'orders': orders, 'services': services})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_part_expense_f(request, expense_id):
    if request.user.job.fin_manager:
        ids = expense_id.split(".")
        id_w = int(ids[1])
        id_p = int(ids[0])
        flag = True
        try:
            par = ids[2]
        except:
            flag = False
        print(id_w, id_p)
        if flag:
            whole_expense = WholeExpense.objects.get(partner__name=par, pk=id_w)
        else:
            whole_expense = WholeExpense.objects.all()[::-1][id_w - 1]
        expense = PartOfLilExpense.objects.filter(whole_expense=whole_expense)[id_p - 1]
        print(expense)
        part = expense.part_of_expense
        if request.method == "POST":
            order = part.service.order
            create_order_invoice(order)
            service = request.POST['service'].split('_')
            serv_part = service[1]
            serv = service[0]

            order = Order.objects.get(number=int(request.POST['number']))
            service = Service.objects.get(name=serv, order=order)
            expense.part_of_expense = PartOfExpense.objects.get(service=service, name=serv_part)
            expense.money = int(request.POST['money'])
            expense.currency = request.POST['currency']
            expense.comments = request.POST['comment']
            expense.save()
            sum = 0
            for exp in part.expenses.all():
                sum += int(exp.money)
            print(sum)
            print(part.expenses.all())
            if sum >= part.money:
                print("!!!!!!!!!!!")
                part.payed = True
            else:
                print("123456")
                part.payed = False
            part.save()
            return redirect('/expenses/')
        orders = Order.objects.all()
        services = Service.objects.filter(order=expense.part_of_expense.service.order)
        return render(request, 'change_part_expense_f.html', {'expense': expense, 'orders': orders, 'services': services})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_part_expense_m(request, expense_id):
    if request.user.job.manager:
        ids = expense_id.split(".")
        id_w = int(ids[1])
        id_p = int(ids[0])
        flag = True
        try:
            par = ids[2]
        except:
            flag = False
        print(id_w, id_p)
        if flag:
            whole_expense = WholeExpense.objects.get(partner__name=par, pk=id_w)
        else:
            whole_expense = WholeExpense.objects.all()[::-1][id_w - 1]
        expense = PartOfLilExpense.objects.filter(whole_expense=whole_expense)[id_p - 1]
        print(expense)
        part = expense.part_of_expense
        if request.method == "POST":
            order = part.service.order
            create_order_invoice(order)
            expense.money = int(request.POST['money'])
            expense.currency = request.POST['currency']
            expense.comments = request.POST['comment']
            expense.save()
            sum = 0
            for exp in part.expenses.all():
                sum += int(exp.money)
            if sum >= part.money:
                part.payed = True
            else:
                part.payed = False
            part.save()
            return redirect('/expenses/')
        orders = Order.objects.all()
        services = Service.objects.filter(order=expense.part_of_expense.service.order)
        return render(request, 'change_part_expense_m.html', {'expense': expense, 'orders': orders, 'services': services})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_part_income_f(request, income_id):
    if request.user.job.fin_manager:
        ids = income_id.split(".")
        id_w = int(ids[1])
        id_p = int(ids[0])
        whole_income = WholeIncome.objects.all()[::-1][id_w - 1]
        income = PartOfIncome.objects.filter(whole_income=whole_income)[id_p - 1]
        if request.method == "POST":
            order = income.order
            create_order_invoice(order)
            income.order = Order.objects.get(number=int(request.POST['number']))
            income.money = int(request.POST['money'])
            income.currency = request.POST['currency']
            income.comments = request.POST['comment']
            income.save()
            return redirect('/incomes/')
        orders = Order.objects.all()
        return render(request, 'change_part_income_f.html', {'income': income, 'orders': orders})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_income_f(request, income_id):
    if request.user.job.fin_manager:
        income = WholeIncome.objects.all()[::-1][int(income_id) - 1]
        if request.method == "POST":
            try:
                income.file = request.FILES['file']
            except:
                pass
            income.created = request.POST['date']
            income.partner = Partner.objects.get(name=request.POST['partner'])
            income.way_to_pay = request.POST['way']
            income.money = int(request.POST['money'])
            income.currency = request.POST['currency']
            income.comments = request.POST['comment']
            income.save()
            return redirect('/incomes/')
        date = str(income.created)
        return render(request, 'change_income_f.html', {'income': income, 'partners': Partner.objects.all(), 'date': date})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_part_income_m(request, income_id):
    if request.user.job.manager:
        ids = income_id.split(" ")
        id_w = int(ids[1])
        id_p = int(ids[0])
        whole_income = WholeIncome.objects.all()[::-1][id_w - 1]
        income = PartOfIncome.objects.filter(whole_income=whole_income)[id_p - 1]
        if request.method == "POST":
            order = income.order
            create_order_invoice(order)
            income.money = int(request.POST['money'])
            income.currency = request.POST['currency']
            income.comments = request.POST['comment']
            income.save()
            return redirect('/incomes/')
        orders = Order.objects.all()
        return render(request, 'change_part_income_m.html', {'income': income, 'orders': orders})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_income_m(request, income_id):
    if request.user.job.manager:
        income = WholeIncome.objects.all()[::-1][int(income_id) - 1]
        if request.method == "POST":
            try:
                income.file = request.FILES['file']
            except:
                pass
            income.created = request.POST['date']
            income.way_to_pay = request.POST['way']
            income.money = int(request.POST['money'])
            income.currency = request.POST['currency']
            income.comments = request.POST['comment']
            income.save()
            return redirect('/incomes/')
        date = str(income.created)
        return render(request, 'change_income_m.html', {'income': income, 'partners': Partner.objects.all(), 'date': date})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_part_income(request, income_id):
    if request.user.job.admin:
        ids = income_id.split(" ")
        id_w = int(ids[1])
        id_p = int(ids[0])
        whole_income = WholeIncome.objects.all()[::-1][id_w - 1]
        income = PartOfIncome.objects.filter(whole_income=whole_income)[id_p - 1]
        if request.method == "POST":
            order = income.order
            create_order_invoice(order)
            try:
                if request.POST['delete'] == 'on':
                    income.delete()
                    return redirect('/incomes/')
            except:
                pass
            income.order = Order.objects.get(number=int(request.POST['number']))
            income.money = int(request.POST['money'])
            income.currency = request.POST['currency']
            income.comments = request.POST['comment']
            income.save()
            return redirect('/incomes/')
        orders = Order.objects.all()
        return render(request, 'change_part_income.html', {'income': income, 'orders': orders})
    else:
        return redirect('/')


@login_required(login_url='/login/')
def change_income(request, income_id):
    if request.user.job.admin:
        income = WholeIncome.objects.all()[::-1][int(income_id) - 1]
        if request.method == "POST":
            try:
                if request.POST['delete'] == 'on':
                    income.delete()
                    return redirect('/incomes/')
            except:
                pass
            try:
                income.file = request.FILES['file']
            except:
                pass
            income.created = request.POST['date']
            income.partner = Partner.objects.get(name = request.POST['partner'])
            income.way_to_pay = request.POST['way']
            income.money = int(request.POST['money'])
            income.currency = request.POST['currency']
            income.comments = request.POST['comment']
            income.save()
            return redirect('/incomes/')
        date = str(income.created)
        return render(request, 'change_income.html', {'income': income, 'partners': Partner.objects.all(), 'date': date})
    else:
        return redirect('/')


def create_order_invoice(order):
    file = []
    file.append(order.number)  # number 0
    file.append(order.last_name)  # name 1
    file.append('')  # addres 2
    file.append(order.partner.vat)  # vat 3
    try:
        file.append(GetPayedDay.objects.get(order=order, percent=100).date)  # date 4
    except:
        file.append('')
    file.append('')  # remarks 5
    file.append([])  # unit 6
    file.append([])  # kolvo 7
    file.append([])  # price 8
    file.append(order.price)  # total 9
    file.append(0)  # charge 10
    file.append([])  # serv 11
    for sr in order.services.all():
        file[6].append(sr.price_for_buyer)
        file[8].append(sr.price_for_buyer)
        file[7].append(1)
        file[11].append(sr.name)
    file.append('')  # data 12
    file.append([])  # res 13
    file.append([])  # eur 14
    file.append([])  # remarks 15
    for inc in order.incomes.all():
        a = str(inc.whole_income.created).split(' ')
        date = str(a[0]).split('-')
        date = ' . '.join(reversed(date))
        file[13].append(date)
        file[14].append(inc.money)
        file[15].append(inc.whole_income.way_to_pay)
    price_per_unit = file[6]
    invoice_order = file[0]
    customer_name = file[1]
    customer_adress = file[2]
    customer_VAT = file[3]
    due_date = file[4]
    remarks = file[5]
    quantity = file[7]
    amount = file[8]
    total_amount = file[9]
    reverse_charge = file[10]
    services = file[11]
    dannie_o_cliente = file[12]
    payments_received = file[13]
    EUR_arr = file[14]
    Remarks_arr = file[15]
    wb = openpyxl.load_workbook(filename='/Users/ksenka/PycharmProjects/Uleima/invoice+order.xlsx') #указываем ПОЛНЫЙ путь к файлу-шаблону
    sheet = wb['Лист1'] #Выбираем нужный лист в данном файле(книге excel)
    date = datetime.strftime(datetime.now(), "%Y.%m.%d(%H%M%S)")
    #png_loc = r'C:\Users\User\Desktop\logo.png' #Сюда вставляется картинка компании
    ws = wb.active
    #my_png = openpyxl.drawing.image.Image(png_loc)
    #ws.add_image(my_png, 'A1')
    sheet['G3'] = invoice_order
    sheet['G4'] = order.created
    sheet['G7'] = customer_name
    sheet['G8'] = customer_adress
    sheet['G9'] = customer_VAT
    sheet['I22'] = due_date
    sheet['A28'] = ("Remarks:" + remarks)
    sheet['G15'] = int(total_amount.split(' ')[0])
    sheet['H15'] = 1
    sheet['I19'] = int(reverse_charge)
    sheet = wb['Лист2']  # Выбираем нужный лист в данном файле(книге excel)
    #date = datetime.strftime(datetime.now(), "%Y%m%d%H%M%S")
    # png_loc = r'C:\Users\User\Desktop\logo.png'  # Сюда вставляется картинка компании
    ws = wb.active
    # my_png = openpyxl.drawing.image.Image(png_loc)
    # ws.add_image(my_png, 'A3')

    sheet['G3'] = invoice_order
    sheet['G7'] = customer_name

    bd = Side(style='thin', color="000000")
    ft = Font(name='Arial', size=14)

    count = 14

    for i in range(len(services)):
        sheet['A' + str(count)] = services[i]
        cell = ws['A' + str(count)]
        cell.font = ft
        count += 1
    count = 14

    for i in range(len(price_per_unit)):
        sheet['G' + str(count)] = int(price_per_unit[i])
        cell = ws['G' + str(count)]
        cell.font = ft
        sheet['G' + str(count)].border = Border(left=bd, right=bd)
        count += 1
    count = 14

    for i in range(len(quantity)):
        sheet['H' + str(count)] = int(quantity[i])
        cell = ws['H' + str(count)]
        cell.font = ft
        sheet['H' + str(count)].border = Border(left=bd, right=bd)
        count += 1
    count = 14

    for i in range(len(amount)):
        sheet['I' + str(count)] = int(amount[i])
        cell = ws['I' + str(count)]
        cell.font = ft
        sheet['I' + str(count)].border = Border(left=bd, right=bd)
        cell = ws['I' + str(count)]
        cell.font = Font(bold=True)
        count += 1

    # count += 2
    # sheet['A' + str(count)] = dannie_o_cliente
    # count += 4

    sheet['A' + str(count)].border = Border(top=bd)
    sheet['B' + str(count)].border = Border(top=bd)
    sheet['C' + str(count)].border = Border(top=bd)
    sheet['D' + str(count)].border = Border(top=bd)
    sheet['E' + str(count)].border = Border(top=bd)
    sheet['F' + str(count)].border = Border(top=bd, left=bd, bottom=bd)
    sheet['F' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['G' + str(count)].border = Border(top=bd, bottom=bd)
    sheet['G' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['H' + str(count)].border = Border(top=bd, bottom=bd, right=bd)
    sheet['H' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['I' + str(count)].border = Border(top=bd, bottom=bd, right=bd, left=bd)
    sheet['I' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")

    sheet['G' + str(count)] = 'TOTAL AMOUNT (EUR):'
    cell = ws['G' + str(count)]
    cell.font = Font(name='Arial', size=14)

    sheet['I' + str(count)] = total_amount
    cell = ws['I' + str(count)]
    cell.font = Font(name='Arial', size=14)
    count += 3
    sheet['A' + str(count)] = 'Payments received'
    sheet['A' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['A' + str(count)].border = Border(top=bd)
    cell = ws['A' + str(count)]
    cell.font = Font(name='Arial', size=14)
    sheet['B' + str(count)] = 'EUR'
    sheet['B' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['B' + str(count)].border = Border(top=bd)
    cell = ws['B' + str(count)]
    cell.font = Font(name='Arial', size=14)
    sheet['C' + str(count)] = 'Remarks'
    sheet['C' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['C' + str(count)].border = Border(top=bd, right=bd)
    cell = ws['C' + str(count)]
    cell.font = Font(name='Arial', size=14)
    count += 1
    count_1 = count

    for i in range(len(payments_received)):
        sheet['A' + str(count)] = payments_received[i]
        cell = ws['A' + str(count)]
        cell.font = ft
        count += 1
    count = count_1

    for i in range(len(EUR_arr)):
        sheet['B' + str(count)] = EUR_arr[i]
        cell = ws['B' + str(count)]
        cell.font = ft
        count += 1
    count = count_1

    for i in range(len(Remarks_arr)):
        sheet['C' + str(count)] = Remarks_arr[i]
        cell = ws['C' + str(count)]
        cell.font = ft
        sheet['C' + str(count)].border = Border(right=bd)
        count += 1
    # count += 1

    sheet['A' + str(count)] = 'Payment due'
    sheet['A' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    cell = ws['A' + str(count)]
    cell.font = Font(name='Arial', size=14, color=colors.RED)
    sheet['A' + str(count)].border = Border(top=bd, bottom=bd)

    sheet['B' + str(count)] = due_date
    sheet['B' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    cell = ws['B' + str(count)]
    cell.font = Font(name='Arial', size=14, color=colors.RED)
    sheet['B' + str(count)].border = Border(top=bd, bottom=bd)

    sheet['C' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['C' + str(count)].border = Border(top=bd, right=bd, bottom=bd)
    wb.save(filename="/Users/ksenka/PycharmProjects/Uleima/media/" + str(invoice_order) + "_" + date + ".xlsx")
    wb.close()
    order.file_base = order.file_base + " " + "media/" + str(invoice_order) + "_" + date + ".xlsx"
    order.save()


def order_in_or(request):
    price_per_unit = request.GET.getlist('price_per_unit[]')
    invoice_order = request.GET.get('invoice_order')
    customer_name = request.GET.get('customer_name')
    customer_adress = request.GET.get('customer_adress')
    customer_VAT = request.GET.get('customer_VAT')
    due_date = request.GET.get('due_date')
    remarks = request.GET.get('remarks')
    quantity = request.GET.getlist('quantity[]')
    amount = request.GET.getlist('amount[]')
    total_amount = request.GET.get('total_amount')
    reverse_charge = request.GET.get('reverse_charge')
    services = request.GET.getlist('services[]')
    dannie_o_cliente = request.GET.get('dannie_o_cliente')
    payments_received = request.GET.getlist('payments_received[]')
    EUR_arr = request.GET.getlist('EUR_arr[]')
    Remarks_arr = request.GET.getlist('Remarks_arr[]')
    created = request.GET.get('created')
    wb = openpyxl.load_workbook(filename='/Users/ksenka/PycharmProjects/Uleima/invoice+order.xlsx') #указываем ПОЛНЫЙ путь к файлу-шаблону
    sheet = wb['Лист1'] #Выбираем нужный лист в данном файле(книге excel)
    date = datetime.strftime(datetime.now(), "%Y%m%d%H%M%S")
    #png_loc = r'C:\Users\User\Desktop\logo.png' #Сюда вставляется картинка компании
    ws = wb.active
    #my_png = openpyxl.drawing.image.Image(png_loc)
    #ws.add_image(my_png, 'A1')
    sheet['G3'] = invoice_order
    sheet['G4'] = created
    sheet['G7'] = customer_name
    sheet['G8'] = customer_adress
    sheet['G9'] = customer_VAT
    sheet['I22'] = due_date
    sheet['A28'] = ("Remarks:" + remarks)
    sheet['G15'] = int(total_amount.split(' ')[0])
    sheet['H15'] = 1
    sheet['I19'] = int(reverse_charge)
    sheet = wb['Лист2']  # Выбираем нужный лист в данном файле(книге excel)
    date = datetime.strftime(datetime.now(), "%Y%m%d%H%M%S")
    # png_loc = r'C:\Users\User\Desktop\logo.png'  # Сюда вставляется картинка компании
    ws = wb.active
    # my_png = openpyxl.drawing.image.Image(png_loc)
    # ws.add_image(my_png, 'A3')

    sheet['G3'] = invoice_order
    sheet['G7'] = customer_name

    bd = Side(style='thin', color="000000")
    ft = Font(name='Arial', size=14)

    count = 14

    for i in range(len(services)):
        sheet['A' + str(count)] = services[i]
        cell = ws['A' + str(count)]
        cell.font = ft
        count += 1
    count = 14

    for i in range(len(price_per_unit)):
        sheet['G' + str(count)] = int(price_per_unit[i])
        cell = ws['G' + str(count)]
        cell.font = ft
        sheet['G' + str(count)].border = Border(left=bd, right=bd)
        count += 1
    count = 14

    for i in range(len(quantity)):
        sheet['H' + str(count)] = int(quantity[i])
        cell = ws['H' + str(count)]
        cell.font = ft
        sheet['H' + str(count)].border = Border(left=bd, right=bd)
        count += 1
    count = 14

    for i in range(len(amount)):
        sheet['I' + str(count)] = int(amount[i])
        cell = ws['I' + str(count)]
        cell.font = ft
        sheet['I' + str(count)].border = Border(left=bd, right=bd)
        cell = ws['I' + str(count)]
        cell.font = Font(bold=True)
        count += 1

    # count += 2
    # sheet['A' + str(count)] = dannie_o_cliente
    # count += 4

    sheet['A' + str(count)].border = Border(top=bd)
    sheet['B' + str(count)].border = Border(top=bd)
    sheet['C' + str(count)].border = Border(top=bd)
    sheet['D' + str(count)].border = Border(top=bd)
    sheet['E' + str(count)].border = Border(top=bd)
    sheet['F' + str(count)].border = Border(top=bd, left=bd, bottom=bd)
    sheet['F' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['G' + str(count)].border = Border(top=bd, bottom=bd)
    sheet['G' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['H' + str(count)].border = Border(top=bd, bottom=bd, right=bd)
    sheet['H' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['I' + str(count)].border = Border(top=bd, bottom=bd, right=bd, left=bd)
    sheet['I' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")

    sheet['G' + str(count)] = 'TOTAL AMOUNT (EUR):'
    cell = ws['G' + str(count)]
    cell.font = Font(name='Arial', size=14)

    sheet['I' + str(count)] = total_amount
    cell = ws['I' + str(count)]
    cell.font = Font(name='Arial', size=14)
    count += 3
    sheet['A' + str(count)] = 'Payments received'
    sheet['A' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['A' + str(count)].border = Border(top=bd)
    cell = ws['A' + str(count)]
    cell.font = Font(name='Arial', size=14)
    sheet['B' + str(count)] = 'EUR'
    sheet['B' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['B' + str(count)].border = Border(top=bd)
    cell = ws['B' + str(count)]
    cell.font = Font(name='Arial', size=14)
    sheet['C' + str(count)] = 'Remarks'
    sheet['C' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['C' + str(count)].border = Border(top=bd, right=bd)
    cell = ws['C' + str(count)]
    cell.font = Font(name='Arial', size=14)
    count += 1
    count_1 = count

    for i in range(len(payments_received)):
        sheet['A' + str(count)] = payments_received[i]
        cell = ws['A' + str(count)]
        cell.font = ft
        count += 1
    count = count_1

    for i in range(len(EUR_arr)):
        sheet['B' + str(count)] = EUR_arr[i]
        cell = ws['B' + str(count)]
        cell.font = ft
        count += 1
    count = count_1

    for i in range(len(Remarks_arr)):
        sheet['C' + str(count)] = Remarks_arr[i]
        cell = ws['C' + str(count)]
        cell.font = ft
        sheet['C' + str(count)].border = Border(right=bd)
        count += 1
    # count += 1

    sheet['A' + str(count)] = 'Payment due'
    sheet['A' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    cell = ws['A' + str(count)]
    cell.font = Font(name='Arial', size=14, color=colors.RED)
    sheet['A' + str(count)].border = Border(top=bd, bottom=bd)

    sheet['B' + str(count)] = due_date
    sheet['B' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    cell = ws['B' + str(count)]
    cell.font = Font(name='Arial', size=14, color=colors.RED)
    sheet['B' + str(count)].border = Border(top=bd, bottom=bd)

    sheet['C' + str(count)].fill = PatternFill(fgColor="E0E0E0", fill_type="solid")
    sheet['C' + str(count)].border = Border(top=bd, right=bd, bottom=bd)
    wb.save(filename="/Users/ksenka/PycharmProjects/Uleima/media/" + str(invoice_order) + "_" + date + ".xlsx")
    wb.close()
    return HttpResponse("media/" + str(invoice_order) + "_" + date + ".xlsx")


def get_check1(request):
    partner = request.GET.get('partner')
    debt = 0
    content = Partner.objects.get(name=partner).incomes.all()
    for order in Partner.objects.get(name=partner).orders.all():
        debt += int(order.price.split(" ")[0])
    for income in content:
        debt -= income.money
    from_date = request.GET.get('from_date')
    if from_date != '':
        from_date = from_date.split('-')
        from_date = dt.date(int(from_date[0]), int(from_date[1]), int(from_date[2]))
    to_date = request.GET.get('to_date')
    if to_date != '':
        to_date = to_date.split('-')
        to_date = dt.date(int(to_date[0]), int(to_date[1]), int(to_date[2]))
    way_s = request.GET.get('way_s')
    if way_s == '-':
        way_s = ''
    number_s = request.GET.get('number_s')
    number_n_s = request.GET.get('number_n_s')
    sum_start = request.GET.get('sum_start').split(' ')
    if sum_start[0] == '':
        sum_start[0] = 0
    name_s = request.GET.get('name_s')
    month_s = request.GET.get('month_s')
    sum_end = request.GET.get('sum_end').split(' ')
    if sum_end[0] == '':
        sum_end[0] = 1000000000
    if len(sum_start) >= 2:
        if sum_start[1] == '':
            sum_start[1] = sum_end[1]
    else:
        if len(sum_end) >= 2:
            sum_start.append(sum_end[1])
        else:
            sum_start.append('')
    content = WholeIncome.objects.filter(
        partner__name=partner,
        way_to_pay__contains=way_s,
        money__gte=int(sum_start[0]),
        money__lte=int(sum_end[0]),
        currency__contains=sum_start[1],
    )
    if from_date == '' and to_date == '':
        pass
    elif from_date == '':
        content = content.filter(
            created__lte=to_date
        )
    elif to_date == '':
        content = content.filter(
            created__gte=from_date
        )
    else:
        content = content.filter(
            created__range=[from_date, to_date],
        )
    if month_s.isnumeric():
        content = content.filter(
            incomes__order__end_month__startswith=month_s
        )
    if number_s != '' or number_n_s != '' or name_s != '' or month_s != '':
        new_content = []
        for one in content:
            if len(one.incomes.all()) != 0:
                filtered = one.incomes.filter(
                    order__number__contains=number_s,
                    order__ref__contains=number_n_s,
                    order__last_name__contains=name_s,
                )
                filtered = filtered.filter(
                    order__end_month__startswith=str(month_s)
                )
                if len(filtered) != 0 and one not in new_content:
                    new_content.append(one)
        content = new_content
    file = [[], [], [], [], [], [], [], [], [], []]
    for one in content:
        a = str(one.created).split(' ')
        date = str(a[0]).split('-')
        date = ' . '.join(reversed(date))
        one.created = date
        file[0].append(one.created)
        file[1].append(one.way_to_pay)
        file[2].append('-')
        file[3].append('-')
        file[4].append(str(one.money) + " " + one.currency)
        file[5].append('-')
        file[6].append('-')
        for lil_income in one.incomes.all():
            file[0].append('')
            file[1].append('')
            file[2].append(lil_income.order.number)
            file[3].append(lil_income.order.ref)
            file[4].append(str(lil_income.money) + " " + lil_income.currency)
            file[5].append(lil_income.order.last_name)
            file[6].append(lil_income.order.end_month)
    return render(request, 'partner_table.html', {'content': content, 'file': file, 'income': True, 'partner': partner, 'e_or_i': 'incomes', 'debt': debt})


def get_check(request):
    partner = request.GET.get('partner')
    debt = 0
    content = Partner.objects.get(name=partner).expenses.all()
    for expense in PartOfExpense.objects.filter(partner=Partner.objects.get(name=partner)):
        debt += expense.money
    for expense in WholeExpense.objects.filter(partner=Partner.objects.get(name=partner)):
        debt -= expense.money
    green = request.GET.get('green')
    red = request.GET.get('red')
    if green == 'true':
        green = True
    else:
        green= False
    if red == 'true':
        red = True
    else:
        red= False
    from_date = request.GET.get('from_date')
    if from_date != '':
        from_date = from_date.split('-')
        from_date = dt.date(int(from_date[0]), int(from_date[1]), int(from_date[2]))
    to_date = request.GET.get('to_date')
    if to_date != '':
        to_date = to_date.split('-')
        to_date = dt.date(int(to_date[0]), int(to_date[1]), int(to_date[2]))
    way_s = request.GET.get('way_s')
    if way_s == '-':
        way_s = ''
    number_s = request.GET.get('number_s')
    number_n_s = request.GET.get('number_n_s')
    service_s = request.GET.get('service_s')
    sum_start = request.GET.get('sum_start').split(' ')
    if sum_start[0] == '':
        sum_start[0] = 0
    left_start = request.GET.get('left_start').split(' ')
    if left_start[0] == '':
        left_start[0] = 0
    name_s = request.GET.get('name_s')
    month_s = request.GET.get('month_s')
    sum_end = request.GET.get('sum_end').split(' ')
    if sum_end[0] == '':
        sum_end[0] = 1000000000
    left_end = request.GET.get('left_end').split(' ')
    if left_end[0] == '':
        left_end[0] = 1000000000
    if len(sum_start) >= 2:
        if sum_start[1] == '':
            sum_start[1] = sum_end[1]
    else:
        if len(sum_end) >= 2:
            sum_start.append(sum_end[1])
        else:
            sum_start.append('')
    file = [[], [], [], [], [], [], [], [], [], []]
    content = PartOfLilExpense.objects.filter(whole_expense__partner__name=partner)
    content = content.filter(
        money__gte=int(sum_start[0]),
        money__lte=int(sum_end[0]),
        currency__contains=sum_start[1],
        whole_expense__way_to_pay__startswith=way_s,
    )
    print("!!!", len(content))
    if from_date == '' and to_date == '':
        pass
    elif from_date == '':
        content = content.filter(
           whole_expense__created__lte=to_date
       )
    elif to_date == '':
        content = content.filter(
           whole_expense__created__gte=from_date
       )
    else:
        content = content.filter(
           whole_expense__created__range=[from_date, to_date],
       )
    print("!!!", len(content))

    content = content.filter(
       part_of_expense__service__order__last_name__contains=name_s,
       part_of_expense__left_to_pay__lte=left_end[0],
       part_of_expense__left_to_pay__gte=left_start[0],
       part_of_expense__service__name__contains=service_s,
       part_of_expense__service__order__number__contains=number_s,
       part_of_expense__service__order__ref__contains=number_n_s,
       part_of_expense__service__order__end_month__startswith=str(month_s),
    )
    print("!!!", len(content))
    if green and red:
        pass
    elif green:
        content = content.filter(
            part_of_expense__payed=True
        )
    elif red:
        content = content.filter(
            part_of_expense__payed=False
        )
    for expense in content:
        a = str(expense.whole_expense.created).split(' ')
        date = str(a[0]).split('-')
        date = ' . '.join(reversed(date))
        expense.whole_expense.created = date
    print("!!!", len(content))
    for lil in content:
        if lil.part_of_expense.payed:
            file[0].append('✅')
        else:
            file[0].append('❌')
        file[1].append(lil.whole_expense.created)
        file[2].append(lil.whole_expense.way_to_pay)
        file[3].append(lil.part_of_expense.order_number)
        file[4].append(lil.part_of_expense.service.order.ref)
        file[5].append(str(lil.part_of_expense.service.name) + " " + str(lil.part_of_expense.name))
        file[6].append(str(lil.money) + " " + str(lil.whole_expense.currency))
        file[7].append(
            str(lil.part_of_expense.left_to_pay) + " " + str(lil.part_of_expense.currency) + " (" + percent(
                lil.part_of_expense.money, lil.part_of_expense.left_to_pay) + "%)")
        file[8].append(lil.part_of_expense.service.order.last_name)
        file[9].append(lil.part_of_expense.service.order.end_month)
    return render(request, 'partner_table.html', {'content': content, 'file': file, 'partner': partner, 'e_or_i': 'expenses', 'debt': debt, 'from_partner_table': False})


def get_order(request):
    number = request.GET.get('number')
    try:
        orders = Order.objects.filter(manager__first_name__startswith=number.split()[0], manager__last_name__startswith=number.split()[1])
        if len(orders) != 0:
            found = True
            for order in orders:
                a = order.created
                date = datetime.strftime(a, "%d. %m. %Y%H%M%S")[:12]
                order.created = date
            return render(request, 'order_result.html', {'orders': orders, 'found': found})
        else:
            found = False
        orders = Order.objects.filter(number__startswith=number)
        if len(orders) != 0:
            found = True
            for order in orders:
                a = order.created
                date = datetime.strftime(a, "%d. %m. %Y%H%M%S")[:12]
                order.created = date
            return render(request, 'order_result.html', {'orders': orders, 'found': found})
        else:
            found = False
        orders = Order.objects.filter(last_name__startswith=number)
        if len(orders) != 0:
            found = True
            for order in orders:
                a = order.created
                date = datetime.strftime(a, "%d. %m. %Y%H%M%S")[:12]
                order.created = date
            return render(request, 'order_result.html', {'orders': orders, 'found': found})
        else:
            found = False
        orders = Order.objects.filter(partner__name__startswith=number)
        if len(orders) != 0:
            found = True
            for order in orders:
                a = order.created
                date = datetime.strftime(a, "%d. %m. %Y%H%M%S")[:12]
                order.created = date
            return render(request, 'order_result.html', {'orders': orders, 'found': found})
        else:
            found = False
    except Exception as e:
        orders = []
        found = False
    return render(request, 'order_result.html', {'orders': orders, 'found': found})


def get_order1(request):
    partner = request.GET.get('partner')
    last_name = request.GET.get('last_name')
    manager = request.GET.get('manager')
    create_date = request.GET.get('create_date')
    tour_date = request.GET.get('tour_date')
    try:
        orders = Order.objects.filter(last_name__contains=last_name, partner__name__contains=partner,
                                      manager__first_name__startswith=manager.split()[0],
                                      manager__last_name__startswith=manager.split()[1])
        if tour_date != '':
            tour_date = tour_date.split('. ')
            date = dt.date(int(tour_date[2]), int(tour_date[1]), int(tour_date[0]))
            orders = orders.filter(tour_starts=date)
        if create_date != '':
            create_date = create_date.split('. ')
            date = dt.date(int(create_date[2]), int(create_date[1]), int(create_date[0]))
            orders = orders.filter(created=date)
        if len(orders) != 0:
            found = True
            for order in orders:
                a = order.created
                date = datetime.strftime(a, "%d. %m. %Y%H%M%S")[:12]
                order.created = date
            return render(request, 'order_result.html', {'orders': orders, 'found': found})
        else:
            found = False
    except:
        orders = []
        found = False
    return render(request, 'order_result.html', {'orders': orders, 'found': found})


@login_required(login_url='/login/')
def search(request):
    orders = Order.objects.all()
    partners = Partner.objects.all()
    managers = User.objects.filter(job__manager=True)
    create_dates = []
    tour_dates = []
    last_names = []
    for order in orders:
        a = order.created
        date = datetime.strftime(a, "%d. %m. %Y%H%M%S")[:12]
        order.created = date
        if date not in create_dates:
            create_dates.append(date)
        a = order.tour_starts
        date = datetime.strftime(a, "%d. %m. %Y%H%M%S")[:12]
        order.tour_starts = date
        if date not in tour_dates:
            tour_dates.append(date)
        if order.last_name not in last_names:
            last_names.append(order.last_name)
    return render(request, 'search.html', {'orders': orders,
                                           'partners': partners,
                                           'managers': managers,
                                           'tour_dates': tour_dates,
                                           'last_names': last_names,
                                           'create_dates': create_dates})


def test(request):
    return render(request, 'test.html')


def logout_view(request):
    logout(request)
    return redirect('/')


@login_required(login_url='/login/')
def index(request):
    return render(request, 'index.html')


def expense_check(request):
    arr = [request.GET.getlist('payeds[]'), request.GET.getlist('dates[]'), request.GET.getlist('ways[]'),
           request.GET.getlist('numbers[]'), request.GET.getlist('refs[]'), request.GET.getlist('services[]'),
           request.GET.getlist('summs[]'), request.GET.getlist('left_to_pays[]'), request.GET.getlist('names[]'),
           request.GET.getlist('months[]')]
    wb = openpyxl.load_workbook(
        filename='/Users/ksenka/PycharmProjects/Uleima/expense_check.xlsx')  # указываем ПОЛНЫЙ путь к файлу-шаблону
    sheet = wb['Лист1']  # Выбираем нужный лист в данном файле(книге excel)
    date = datetime.strftime(datetime.now(), "%Y%m%d%H%M%S")
    count = 2
    _count = count
    for i in range(len(arr[0])):
        sheet['A' + str(_count)] = arr[0][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['B' + str(_count)] = arr[1][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['C' + str(_count)] = arr[2][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['D' + str(_count)] = arr[3][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['E' + str(_count)] = arr[4][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['F' + str(_count)] = arr[5][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['G' + str(_count)] = arr[6][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['H' + str(_count)] = arr[7][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['I' + str(_count)] = arr[8][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['J' + str(_count)] = arr[9][i]
        _count += 1
    wb.save(filename="/Users/ksenka/PycharmProjects/Uleima/media/" + request.GET.get(
        'partner') + "_check_" + date + ".xlsx")
    wb.close()
    return HttpResponse("media/" + request.GET.get('partner') + "_check_" + date + ".xlsx")


def income_check(request):
    arr = [request.GET.getlist('dates[]'), request.GET.getlist('ways[]'),
           request.GET.getlist('numbers[]'), request.GET.getlist('refs[]'), request.GET.getlist('summs[]'),
           request.GET.getlist('names[]'), request.GET.getlist('months[]')]
    wb = openpyxl.load_workbook(
        filename='/Users/ksenka/PycharmProjects/Uleima/income_check.xlsx')  # указываем ПОЛНЫЙ путь к файлу-шаблону
    sheet = wb['Лист1']  # Выбираем нужный лист в данном файле(книге excel)
    date = datetime.strftime(datetime.now(), "%Y%m%d%H%M%S")
    count = 2
    _count = count
    for i in range(len(arr[0])):
        sheet['A' + str(_count)] = arr[0][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['B' + str(_count)] = arr[1][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['C' + str(_count)] = arr[2][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['D' + str(_count)] = arr[3][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['E' + str(_count)] = arr[4][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['F' + str(_count)] = arr[5][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['G' + str(_count)] = arr[6][i]
        _count += 1
    wb.save(filename="/Users/ksenka/PycharmProjects/Uleima/media/" + request.GET.get(
        'partner') + "_check_" + date + ".xlsx")
    wb.close()
    print("media/" + request.GET.get('partner') + "_check_" + date + ".xlsx")
    return HttpResponse("media/" + request.GET.get('partner') + "_check_" + date + ".xlsx")


def income_cassa_file(request):
    arr = [request.GET.getlist('dates[]'), request.GET.getlist('partners[]'), request.GET.getlist('ways[]'),
           request.GET.getlist('numbers[]'), request.GET.getlist('refs[]'), request.GET.getlist('summs[]'),
           request.GET.getlist('names[]'), request.GET.getlist('months[]')]
    wb = openpyxl.load_workbook(
        filename='/Users/ksenka/PycharmProjects/Uleima/income_cassa.xlsx')  # указываем ПОЛНЫЙ путь к файлу-шаблону
    sheet = wb['Лист1']  # Выбираем нужный лист в данном файле(книге excel)
    date = datetime.strftime(datetime.now(), "%Y%m%d%H%M%S")
    count = 2
    _count = count
    for i in range(len(arr[0])):
        sheet['A' + str(_count)] = arr[0][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['B' + str(_count)] = arr[1][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['C' + str(_count)] = arr[2][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['D' + str(_count)] = arr[3][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['E' + str(_count)] = arr[4][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['F' + str(_count)] = arr[5][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['G' + str(_count)] = arr[6][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['H' + str(_count)] = arr[7][i]
        _count += 1
    wb.save(filename="/Users/ksenka/PycharmProjects/Uleima/media/" + "income_cassa" + "_" + date + ".xlsx")
    wb.close()
    return HttpResponse("media/" + "income_cassa" + "_" + date + ".xlsx")


def expense_cassa_file(request):
    arr = [request.GET.getlist('dates[]'), request.GET.getlist('partners[]'), request.GET.getlist('ways[]'),
           request.GET.getlist('numbers[]'), request.GET.getlist('refs[]'), request.GET.getlist('services[]'),
           request.GET.getlist('summs[]'), request.GET.getlist('names[]'), request.GET.getlist('months[]')]
    wb = openpyxl.load_workbook(
        filename='/Users/ksenka/PycharmProjects/Uleima/expense_cassa.xlsx')  # указываем ПОЛНЫЙ путь к файлу-шаблону
    sheet = wb['Лист1']  # Выбираем нужный лист в данном файле(книге excel)
    date = datetime.strftime(datetime.now(), "%Y%m%d%H%M%S")
    count = 2
    _count = count
    for i in range(len(arr[0])):
        sheet['A' + str(_count)] = arr[0][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['B' + str(_count)] = arr[1][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['C' + str(_count)] = arr[2][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['D' + str(_count)] = arr[3][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['E' + str(_count)] = arr[4][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['F' + str(_count)] = arr[5][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['G' + str(_count)] = arr[6][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['H' + str(_count)] = arr[7][i]
        _count += 1
    _count = count
    for i in range(len(arr[0])):
        sheet['I' + str(_count)] = arr[8][i]
        _count += 1
    wb.save(filename="/Users/ksenka/PycharmProjects/Uleima/media/" + "expense_cassa" + "_" + date + ".xlsx")
    wb.close()
    return HttpResponse("media/" + "expense_cassa" + "_" + date + ".xlsx")


def percent(price, payed):
    percent = str((int(payed) / int(price)) * 100)
    percent = percent.split('.')
    return percent[0] + '.' + percent[1][:1]


def download_file(request, filename):
    the_file = 'media/' + filename
    filename = os.path.basename(the_file)
    chunk_size = 8192
    response = StreamingHttpResponse(FileWrapper(open(the_file, 'rb'), chunk_size),
                                     content_type=mimetypes.guess_type(the_file)[0])
    response['Content-Length'] = os.path.getsize(the_file)
    response['Content-Disposition'] = "attachment; filename=%s" % filename
    return response


@login_required(login_url='/login/')
def load_expenses(request):
    number = request.GET.get('number')
    names = []
    for service in Service.objects.filter(order=Order.objects.get(number=number)):
        for expense in service.expense.all():
            if not expense.payed:
                names.append(expense)
    return render(request, 'service_dropdown.html', {'services': names})


@login_required(login_url='/login/')
def load_expenses1(request):
    number = request.GET.get('number')
    names = []
    for service in Service.objects.filter(order=Order.objects.get(number=number)):
        for expense in service.expense.all():
            names.append(expense)
    return render(request, 'service1.html', {'services': names})


@login_required(login_url='/login/')
def load_partner(request):
    partner = request.GET.get('partner')
    e_or_i = request.GET.get('or')
    debt = 0
    if partner == '-----------':
        content = []
    else:
        if e_or_i == 'incomes':
            file = [[], [], [], [], [], [], []]
            content = Partner.objects.get(name=partner).incomes.all()
            for order in Partner.objects.get(name=partner).orders.all():
                debt += int(order.price.split(" ")[0])
            for income in content:
                debt -= income.money
            for index, expense in enumerate(content, start=0):
                a = str(expense.created).split(' ')
                date = str(a[0]).split('-')
                date = ' . '.join(reversed(date))
                expense.created = date
            for one in content:
                file[0].append(one.created)
                file[1].append(one.way_to_pay)
                file[2].append('-')
                file[3].append('-')
                file[4].append(str(one.money) + " " + one.currency)
                file[5].append('-')
                file[6].append('-')
                for lil_income in one.incomes.all():
                    file[0].append('')
                    file[1].append('')
                    file[2].append(lil_income.order.number)
                    file[3].append(lil_income.order.ref)
                    file[4].append(str(lil_income.money) + " " + lil_income.currency)
                    file[5].append(lil_income.order.last_name)
                    file[6].append(lil_income.order.end_month)
        else:
            file = [[], [], [], [], [], [], [], [], [], []]
            content = Partner.objects.get(name=partner).expenses.all()
            for expense in PartOfExpense.objects.filter(partner=Partner.objects.get(name=partner)):
                debt += expense.money
            for expense in WholeExpense.objects.filter(partner=Partner.objects.get(name=partner)):
                debt -= expense.money
            for index, expense in enumerate(content, start=0):
                a = str(expense.created).split(' ')
                date = str(a[0]).split('-')
                date = ' . '.join(reversed(date))
                expense.created = date
            for one in content:
                for lil in one.expenses.all():
                    if lil.part_of_expense.payed:
                        file[0].append('✅')
                    else:
                        file[0].append('❌')
                    file[1].append(one.created)
                    file[2].append(one.way_to_pay)
                    file[3].append(lil.part_of_expense.order_number)
                    file[4].append(lil.part_of_expense.service.order.ref)
                    file[5].append(str(lil.part_of_expense.service.name) + " " + str(lil.part_of_expense.name))
                    file[6].append(str(lil.money) + " " + str(lil.currency))
                    file[7].append(
                        str(lil.part_of_expense.left_to_pay) + " " + str(lil.part_of_expense.currency) + " (" + percent(
                            lil.part_of_expense.money, lil.part_of_expense.left_to_pay) + "%)")
                    file[8].append(lil.part_of_expense.service.order.last_name)
                    file[9].append(lil.part_of_expense.service.order.end_month)
    return render(request, 'partner_table.html',
                  {'content': content, 'e_or_i': e_or_i, 'debt': debt, 'partner': partner, 'file': file, 'prev': request.GET.get('prev'), 'from_partner_table': True})


@login_required(login_url='/login/')
def check(request):
    return render(request, 'check.html', {'partners': Partner.objects.all(), 'prev': 'btn_download', 'content': [], 'e_or_i': 'other', 'file': [[],[],[],[],[],[],[],[],[],[],[],[]]})


@login_required(login_url='/login/')
def new_order(request):
    if request.method == "POST":
        form = OrderForm(request.POST)
        number = Order.objects.all()[::-1][0].number + 1
        count_service = ((len(request.POST) - 14) // 2)
        if form.is_valid():
            created = form.cleaned_data['created']
            manager = User.objects.get(first_name__startswith=form.cleaned_data['manager'].split()[0], last_name__startswith=form.cleaned_data['manager'].split()[1])
            partner = Partner.objects.get(name=form.cleaned_data['partner'])
            ref = form.cleaned_data['ref']
            another_manager = form.cleaned_data['another_manager']
            last_name = form.cleaned_data['last_name']
            end_month = form.cleaned_data['end_month']
            comments = form.cleaned_data['comments']
            tour_starts = form.cleaned_data['tour_starts']
            order = Order.objects.create(number=number, manager=manager, partner=partner, ref=ref,
                                         another_manager=another_manager, last_name=last_name, end_month=end_month,
                                         comments=comments, tour_starts=tour_starts)
            order.save()
            order.created = created
            order.save()
            a = str(order.created).split(' ')
            date = str(a[0]).split('-')
            date = ' . '.join(reversed(date))
            price = 0
            for i in range(count_service):
                service = Service.objects.create(name=request.POST['name_' + str(int(i) + 1)],
                                                 price_for_buyer=int(request.POST['money_' + str(int(i) + 1)]),
                                                 order=Order.objects.get(number=number),
                                                 real_price=int(request.POST['money_' + str(int(i) + 1)]))
                price += int(request.POST['money_' + str(int(i) + 1)])
                service.save()
            order.price = price
            order.save()
        return redirect('/')
    else:
        form = OrderForm
        return render(request, 'order_create.html', {'form': form})


@login_required(login_url='/login/')
def order_screen(request, order_id):
    if request.POST:
        if 'percent_gr_1' in request.POST.keys():
            order_gr = Order.objects.get(number=order_id)
            order_gr.get_payed_days.all().delete()
            count_gr = len(request.POST) // 2
            for i in range(count_gr):
                percent = str(request.POST['percent_gr_' + str(int(i) + 1)]) + " "
                date = str(request.POST['date_gr_' + str(int(i) + 1)]) + " "
                date = date.split('-')
                date = dt.date(year=int(date[0]), month=int(date[1]), day=int(date[2]))
                pay_day = GetPayedDay.objects.create(date=date, order=order_gr, percent=percent)
                pay_day.save()
                pay_day.date = date
                pay_day.save()
        elif 'price_b' in request.POST.keys():
            name = request.POST['name']
            price = request.POST['price']
            price_b = request.POST['price_b']
            service = Service.objects.create(order=Order.objects.get(number=order_id), real_price=price, price_for_buyer=price_b, name=name)
            service.save()
        else:
            count_service = ((len(request.POST) - 8) // 2)
            name = request.POST['name']
            service = request.POST['service']
            money = request.POST['money']
            currency = request.POST['currency']
            partner = request.POST['partner']
            comments = request.POST['comments']
            number = request.POST['number']
            expense = PartOfExpense.objects.create(name=name, service=Service.objects.get(name=service, order=Order.objects.get(number=number)), money=money, currency=currency,
                                                   partner=Partner.objects.get(name=partner), comments=comments, left_to_pay=money,
                                                   order_number=number)
            expense.save()
            for i in range(count_service):
                percent = str(request.POST['percent_' + str(int(i) + 1)]) + " "
                date = str(request.POST['date_' + str(int(i) + 1)]) + " "
                date = date.split('-')
                date = dt.date(year=int(date[0]), month=int(date[1]), day=int(date[2]))
                pay_day = PayDay.objects.create(date=date, order=Order.objects.get(number=number), part_of_expense=expense, percent=percent)
                pay_day.save()
                pay_day.date = date
                pay_day.save()
    order = get_object_or_404(Order, number=order_id)
    a = order.created
    date = datetime.strftime(a, "%d. %m. %Y%H%M%S")[:12]
    order.created = date
    sum_buyer = 0
    real_sum = 0
    for service in order.services.all():
        sum_buyer += service.price_for_buyer
        real_sum += service.real_price
    got = sum_buyer - real_sum
    expenses = PartOfExpense.objects.filter(order_number=order.number)
    incomes = order.incomes.all()
    lil_expenses = PartOfLilExpense.objects.filter(
        part_of_expense__order_number=order.number
    )
    try:
        percent = str(got * 100 / sum_buyer)[:5]
    except:
        percent = 0

    for income in incomes:
        a = str(income.whole_income.created).split(' ')
        date = str(a[0]).split('-')
        date = ' . '.join(reversed(date))
        income.whole_income.created = date
    for lili in lil_expenses:
        a = str(lili.whole_expense.created).split(' ')
        print(a)
        date = str(a[0]).split('-')
        date = ' . '.join(reversed(date))
        lili.whole_expense.created = date
        print(lili.whole_expense.created)
    today = datetime.today()
    today = str(today.year) + '-' + str(today.month) + '-' + str(today.day)
    file = []
    file.append(order_id)  # number 0
    file.append(order.last_name)  # name 1
    file.append('')  # adress 2
    file.append(order.partner.vat)  # vat 3
    try:
        file.append(GetPayedDay.objects.get(order=order, percent=100).date)  # date 4
    except:
        file.append('')
    file.append('')  # remarks 5
    file.append([])  # unit 6
    file.append([])  # kolvo 7
    file.append([])  # price 8
    file.append(order.price)  # total 9
    file.append(0)  # charge 10
    file.append([])  # serv 11
    for sr in order.services.all():
        file[6].append(sr.price_for_buyer)
        file[8].append(sr.price_for_buyer)
        file[7].append(1)
        file[11].append(sr.name)
    file.append('')  # data 12
    file.append([])  # res 13
    file.append([])  # eur 14
    file.append([])  # remarks 15
    file.append(order.created)
    for inc in order.incomes.all():
        a = str(inc.whole_income.created).split(' ')
        date = str(a[0]).split('-')
        date = ' . '.join(reversed(date))
        file[13].append(date)
        file[14].append(inc.money)
        file[15].append(inc.whole_income.way_to_pay)
    history = order.file_base.split("media/")[1:]
    context = {
        'order': order,
        'real_sum': real_sum,
        'sum_buyer': sum_buyer,
        'got': got,
        'percent': percent,
        'expenses': expenses,
        'incomes': incomes,
        'form': CreateExpenseForm(number=order_id),
        'today': today,
        'array': file,
        'history': history,
    }
    return render(request, 'order.html', context)


@login_required(login_url='/login/')
def incomes(request):
    if request.method == 'POST':

        if request.POST['num'] != '':
            form = PartOfIncomeForm(request.POST)
            if form.is_valid():
                number = form.cleaned_data['number']
                money = form.cleaned_data['money']
                currency = form.cleaned_data['currency']
                inc_id = request.POST['num']
                comments = form.cleaned_data['comments']
                lil_income = PartOfIncome.objects.create(money=money, currency=currency,
                                                         whole_income=WholeIncome.objects.all()[::-1][int(inc_id) - 1],
                                                         order=Order.objects.get(number=int(number)), comments=comments)
                lil_income.save()
            incomes = WholeIncome.objects.all()
            for income in incomes:
                a = str(income.created).split(' ')
                date = str(a[0]).split('-')
                date = ' . '.join(reversed(date))
                income.created = date
        else:
            myfile = request.FILES['file']
            form = WholeIncomeForm(request.POST, request.FILES)
            if form.is_valid():
                partner = Partner.objects.get(name=form.cleaned_data['partner'])
                money = form.cleaned_data['money']
                way_to_pay = form.cleaned_data['way_to_pay']
                currency = form.cleaned_data['currency']
                comments = form.cleaned_data['comments']
                whole_income = WholeIncome.objects.create(partner=partner, money=money, way_to_pay=way_to_pay,
                                                          currency=currency, file=myfile, comments=comments)
                whole_income.save()
    incomes = WholeIncome.objects.all()
    color = {}
    file = [[], [], [], [], [], [], [], []]
    for index, income in enumerate(incomes, start=0):
        a = str(income.created).split(' ')
        date = str(a[0]).split('-')
        date = ' . '.join(reversed(date))
        income.created = date
        file[0].append(str(date))
        file[1].append(income.partner.name)
        file[2].append(income.way_to_pay)
        file[3].append('')
        file[4].append('')
        file[5].append(str(income.money) + " " + str(income.currency))
        file[6].append('')
        file[7].append('')
        sum = 0
        for lil in income.incomes.all():
            sum += int(lil.money)
            file[0].append('')
            file[1].append('')
            file[2].append('')
            file[3].append(lil.order.number)
            file[4].append(str(lil.order.ref))
            file[5].append(str(lil.money) + " " + str(lil.currency))
            file[6].append(lil.order.last_name)
            file[7].append(lil.order.end_month)
        if sum == int(income.money) or sum == 0:
            color[len(incomes) - index - 1] = "ok"
        else:
            color[len(incomes) - index - 1] = 'not_ok'
    return render(request, 'incomes.html',
                  {'incomes': incomes[::-1], 'form': WholeIncomeForm, 'income_form': PartOfIncomeForm,
                   'row_color': color, 'array': file})


@login_required(login_url='/login/')
def expenses(request):
    if request.method == 'POST':
        if request.POST['num'] != '':
            form = PartOfExpenseForm(request.POST)
            if form.is_valid():
                money = form.cleaned_data['money']
                currency = form.cleaned_data['currency']
                inc_id = request.POST['num']
                service = form.cleaned_data['service']
                number = form.cleaned_data['number']
                comments = form.cleaned_data['comments']
                lil_expense = PartOfExpense.objects.get(name=service, order_number=number)
                if money >= lil_expense.left_to_pay:
                    lil_expense.left_to_pay = 0
                    lil_expense.payed = True
                else:
                    lil_expense.left_to_pay = lil_expense.left_to_pay - money
                lil_expense.save()
                very_lil_expense = PartOfLilExpense.objects.create(money=money, currency=currency,
                                                                   part_of_expense=lil_expense,
                                                                   whole_expense=WholeExpense.objects.all()[::-1][
                                                                       int(inc_id) - 1],
                                                                   comments=comments)
                very_lil_expense.save()
        else:
            myfile = request.FILES['file']
            form = WholeIncomeForm(request.POST, request.FILES)
            if form.is_valid():
                partner = Partner.objects.get(name=form.cleaned_data['partner'])
                money = form.cleaned_data['money']
                way_to_pay = form.cleaned_data['way_to_pay']
                currency = form.cleaned_data['currency']
                comments = form.cleaned_data['comments']
                whole_expense = WholeExpense.objects.create(partner=partner, money=money, way_to_pay=way_to_pay,
                                                            currency=currency, file=myfile, comments=comments)
                whole_expense.save()
    expenses = WholeExpense.objects.all()
    color = {}
    file = [[], [], [], [], [], [], [], [], []]
    for index, expense in enumerate(expenses, start=0):
        a = str(expense.created).split(' ')
        date = str(a[0]).split('-')
        date = ' . '.join(reversed(date))
        expense.created = date
        file[0].append(str(date))
        file[1].append(expense.partner.name)
        file[2].append(expense.way_to_pay)
        file[3].append('-')
        file[4].append('-')
        file[5].append(str(expense.money) + " " + str(expense.currency))
        file[6].append('-')
        file[7].append('-')
        file[8].append('-')
        sum = 0
        for lil in expense.expenses.all():
            sum += int(lil.money)
            file[0].append('')
            file[1].append('')
            file[2].append('')
            file[3].append(lil.part_of_expense.service.order.number)
            file[4].append(str(lil.part_of_expense.service.order.ref))
            file[5].append(str(lil.money) + " " + str(lil.currency))
            file[6].append(lil.part_of_expense.service.order.last_name)
            file[7].append(lil.part_of_expense.service.order.end_month)
            file[8].append(lil.part_of_expense.service.name + "(" + lil.part_of_expense.name + ")")
        if sum == int(expense.money) or sum == 0:
            color[len(expenses) - index - 1] = "ok"
        else:
            color[len(expenses) - index - 1] = 'not_ok'
    return render(request, 'expenses.html',
                  {'incomes': expenses[::-1], 'form': WholeIncomeForm, 'income_form': PartOfExpenseForm,
                   'row_color': color, 'array': file})


@login_required(login_url='/login/')
def order(request):
    return render(request, 'order.html')


@login_required(login_url='/login/')
def calendula(request):
    date = get_new_date(request.GET.get('or')).split('.')
    c = calendar.Calendar(0)
    listislav = c.monthdayscalendar(int(date[1]), int(date[0]))
    days = PayDay.objects.filter(date__month=int(date[0]), date__year=int(date[1])).order_by('date')
    payed_pays = GetPayedDay.objects.filter(date__month=int(date[0]), date__year=int(date[1])).order_by('date')
    lil_kek = [[[], [], [], [], [], [], []] for i in listislav]
    payed_dates = [[[], [], [], [], [], [], []] for i in listislav]
    current_day = 0
    for i in range(len(listislav)):
        for j in range(7):
            if listislav[i][j] != 0:
                current_day += 1
            for dayy in days:
                if int(dayy.date.day) == current_day:
                    lil_kek[i][j].append("Должно быть переведено " + dayy.part_of_expense.partner.name + " " + str(dayy.part_of_expense.money * dayy.percent * 0.01) + " " + dayy.part_of_expense.currency + " по заказу № " + str(dayy.part_of_expense.order_number))
            for dday in payed_pays:
                print(dday.order.price)
                if int(dday.date.day) == current_day:
                    payed_dates[i][j].append("Должно быть оплачено от " + dday.order.partner.name + " " + str(int(dday.order.price.split(' ')[0]) * dday.percent * 0.01) + " " + dday.order.price.split(' ')[1] + " по заказу № " + str(dday.order.number))
    return render(request, 'calendula.html', {'table': listislav, 'dates': lil_kek, 'payed_dates': payed_dates})


def calendula1(request):
    return HttpResponse(get_new_date(request.GET.get('or')))


def my_calendar(request):
    today = str(datetime.now()).split('-')[:2]
    return render(request, 'calendar.html', {'ye': today})


def get_new_date(date):
    date = date.split('.')
    month = int(date[0]) + 1
    if month == 13:
        month = 1
        date[1] = int(date[1])
        date[1] += 1
    return str(month) + '.' + str(date[1])


def now_calendula(request):
    date = get_now_date(request.GET.get('or')).split('.')
    c = calendar.Calendar(0)
    listislav = c.monthdayscalendar(int(date[1]), int(date[0]))
    days = PayDay.objects.filter(date__month=datetime.today().month, date__year=datetime.today().year).order_by(
        'date')
    payed_pays = GetPayedDay.objects.filter(date__month=datetime.today().month, date__year=datetime.today().year).order_by(
        'date')
    lil_kek = [[[], [], [], [], [], [], []] for i in listislav]
    payed_dates = [[[], [], [], [], [], [], []] for i in listislav]
    current_day = 0
    for i in range(len(listislav)):
        for j in range(7):
            if listislav[i][j] != 0:
                current_day += 1
            for dayy in days:
                if int(dayy.date.day) == current_day:
                    lil_kek[i][j].append("Должно быть переведено " + dayy.part_of_expense.partner.name + " " + str(dayy.part_of_expense.money * dayy.percent * 0.01) + " " + dayy.part_of_expense.currency + " по заказу № " + str(dayy.part_of_expense.order_number))
            for dday in payed_pays:
                print(dday.order.price)
                if int(dday.date.day) == current_day:
                    payed_dates[i][j].append("Должно быть оплачено от " + dday.order.partner.name + " " + str(int(dday.order.price.split(' ')[0]) * dday.percent * 0.01) + " " + dday.order.price.split(' ')[1] + " по заказу № " + str(dday.order.number))
    return render(request, 'calendula.html', {'table': listislav, 'dates': lil_kek, 'payed_dates': payed_dates})


def now_calendula1(request):
    return HttpResponse(get_now_date(request.GET.get('or')))


def get_now_date(date):
    date = date.split('.')
    month = int(date[0])
    return str(month) + '.' + date[1]


def old_calendula(request):
    date = get_old_date(request.GET.get('or')).split('.')
    c = calendar.Calendar(0)
    listislav = c.monthdayscalendar(int(date[1]), int(date[0]))
    days = PayDay.objects.filter(date__month=int(date[0]), date__year=int(date[1])).order_by('date')
    payed_pays = GetPayedDay.objects.filter(date__month=int(date[0]), date__year=int(date[1])).order_by('date')
    lil_kek = [[[], [], [], [], [], [], []] for i in listislav]
    payed_dates = [[[], [], [], [], [], [], []] for i in listislav]
    current_day = 0
    for i in range(len(listislav)):
        for j in range(7):
            if listislav[i][j] != 0:
                current_day += 1
            for dayy in days:
                if int(dayy.date.day) == current_day:
                    lil_kek[i][j].append("Должно быть переведено " + dayy.part_of_expense.partner.name + " " + str(dayy.part_of_expense.money * dayy.percent * 0.01) + " " + dayy.part_of_expense.currency + " по заказу № " + str(dayy.part_of_expense.order_number))
            for dday in payed_pays:
                print(dday.order.price)
                if int(dday.date.day) == current_day:
                    payed_dates[i][j].append("Должно быть оплачено от " + dday.order.partner.name + " " + str(int(dday.order.price.split(' ')[0]) * dday.percent * 0.01) + " " + dday.order.price.split(' ')[1] + " по заказу № " + str(dday.order.number))
    return render(request, 'calendula.html', {'table': listislav, 'dates': lil_kek, 'payed_dates': payed_dates})


def old_calendula1(request):
    return HttpResponse(get_old_date(request.GET.get('or')))


def get_old_date(date):
    date = date.split('.')
    month = int(date[0]) - 1
    if month == 0:
        month = 12
        date[1] = int(date[1])
        date[1] -= 1
    return str(month) + '.' + str(date[1])
